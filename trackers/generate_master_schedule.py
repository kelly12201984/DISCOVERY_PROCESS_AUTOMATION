"""
Master Schedule Generator (replaces April's MASTER SCHEDULE.xlsx)
Produces a week-by-week Gantt-style view of all active jobs,
showing which phase each job is in during each week.

Built from ACTUAL operation progress data — not Sched_Start/Sched_End
(which are empty in JobBOSS).  Projects future weeks using a
capacity-constrained simulation: each phase has a fixed weekly hour
budget shared across ALL jobs, so work naturally staggers through
the shop like it does in real life.
"""
import pandas as pd
import os
from datetime import datetime, timedelta

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'JOBBOSS_DATA')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'generated_trackers')

# Total shop hours available PER PHASE PER WEEK (shared across all jobs).
# Tune these to match actual crew sizes / bay availability.
PHASE_CAPACITY_PER_WEEK = {
    'PRE-FAB':      400,   # ~8-10 people cutting/rolling/building parts
    'FABRICATION':  500,   # ~10-12 welders/fitters — biggest crew
    'TESTING':      160,   # ~2-3 people, inspectors + QC
    'CLEANING':     160,   # ~2-3 people, sandblast crew
    'PAINT':        160,   # ~2-3 people, paint crew
    'SHIPPING':     120,   # ~2 people, load-out crew
}

# Phases in production order, with keywords that map operations to phases
PHASES = [
    ('PRE-FAB',     ['Receive Material', 'UT Top', 'UT Bottom', 'UT Shell',
                     'Cut Parts', 'Cut Nozzle', 'Clean & Prep', 'Roll Shell',
                     'Roll Pipe', 'Roll Repad', 'Roll Rim', 'Build Nozzle',
                     'Make Manway', 'Build Hatch', 'Build Door', 'Bld Skirt',
                     'Bld Lift Lugs', 'Bld Grd Lugs', 'B&C parts',
                     'BI, SQ', 'Fab Misc', 'Build Support', 'Build Top',
                     'Build Bottom', 'Misc Fabrication', 'Drill', 'Punch',
                     'Build Flange']),
    ('FABRICATION', ['S/U and Build', 'Set Nozzle', 'Weld Shell', 'Weld Nozzle',
                     'Weld top', 'Weld bottom', 'Weld all', 'Weld ring',
                     'Weld Skirt', 'Weld Shell Nozz',
                     'Install Nozzle', 'Install Internal', 'Install Support',
                     'Install Insul', 'Install Vac', 'Install Baffle',
                     'Install Body Flg', 'Install Bottom Stiff',
                     'Install Top Stiff', 'Install Lift Lugs',
                     'Fit Top', 'Fit Bottom', 'Fit Ladder', 'Fit Handrail',
                     'Fit Platform', 'I&W Bottom', 'I&W Top',
                     'L/O Top', 'L/O Bottom', 'Move Tank',
                     'Tack-Weld', 'Roll Shell, Tack']),
    ('TESTING',     ['Test Tank', 'HYDRO', 'Dye Pen', 'ASME Inspection',
                     'QC Inspection - Test', 'QC Inspection - Internal',
                     'QC Inspection -Top', 'QC Inspection']),
    ('CLEANING',    ['Clean Tank', 'Prepare Tank', 'Move to Paint',
                     'QC Inspection - Sandblast', 'Sandblast']),
    ('PAINT',       ['Blast & Paint', 'Paint', 'QC Inspection - Paint']),
    ('SHIPPING',    ['Load on Truck', 'LOAD ON TRUCK', 'Check BOM', 'Pack Parts',
                     'QC Inspection - Final', 'Prepare Tank for Shipment']),
]

PHASE_NAMES = [p[0] for p in PHASES]


def classify_operation(desc):
    """Map an operation description to a phase name."""
    desc_lower = str(desc).lower()
    for phase_name, keywords in PHASES:
        if any(kw.lower() in desc_lower for kw in keywords):
            return phase_name
    return None


def analyze_job(job_ops):
    """Compute per-phase hours (estimated, actual, remaining) and current phase."""
    if job_ops.empty:
        return None

    sorted_ops = job_ops.sort_values('Sequence')

    # Accumulate hours by phase
    phase_est = {p: 0.0 for p in PHASE_NAMES}
    phase_remaining = {p: 0.0 for p in PHASE_NAMES}
    phase_has_started = {p: False for p in PHASE_NAMES}
    phase_all_done = {p: True for p in PHASE_NAMES}

    current_phase = None

    for _, op in sorted_ops.iterrows():
        est = op['Est_Total_Hrs'] or 0
        pct = (op['Run_Pct_Complete'] or 0) / 100.0
        status = op['Status']
        remaining = est * (1 - pct)

        phase = classify_operation(op.get('Description', ''))
        if phase is None:
            # Assign unclassified ops based on sequence position
            total = len(sorted_ops)
            seq_pct = op['Sequence'] / max(total - 1, 1) if total > 1 else 0
            if seq_pct < 0.25:
                phase = 'PRE-FAB'
            elif seq_pct < 0.65:
                phase = 'FABRICATION'
            elif seq_pct < 0.80:
                phase = 'TESTING'
            elif seq_pct < 0.90:
                phase = 'CLEANING'
            else:
                phase = 'SHIPPING'

        phase_est[phase] += est
        phase_remaining[phase] += remaining

        if status in ('S', 'C') or pct > 0:
            phase_has_started[phase] = True
        if status != 'C' and pct < 100:
            phase_all_done[phase] = False

    # Determine current phase: latest phase that has started but isn't fully done
    for phase in PHASE_NAMES:
        if phase_has_started[phase] and not phase_all_done[phase]:
            current_phase = phase
    # If nothing in progress, find earliest phase with remaining hours
    if current_phase is None:
        for phase in PHASE_NAMES:
            if phase_remaining[phase] > 0:
                current_phase = phase
                break

    total_est = sum(phase_est.values())
    total_remaining = sum(phase_remaining.values())

    return {
        'phase_est': phase_est,
        'phase_remaining': phase_remaining,
        'current_phase': current_phase or 'UNKNOWN',
        'total_est_hrs': total_est,
        'total_remaining_hrs': total_remaining,
        'pct_complete': ((total_est - total_remaining) / total_est * 100) if total_est > 0 else 0,
    }


def schedule_all_jobs(job_analyses, today, num_weeks=24):
    """Capacity-constrained scheduler: simulates week-by-week flow through the shop.

    Instead of projecting each job independently (which piles all similar jobs
    into the same phase at the same time), this allocates a shared weekly hour
    budget per phase across all competing jobs.  Jobs are prioritized by
    promised date, then by how far along they already are.

    Returns: {job_key: {week_date: phase_name}} for every job/week with work.
    """
    snap_to_monday = lambda d: d - timedelta(days=d.weekday())
    start_week = snap_to_monday(today)
    week_dates = [start_week + timedelta(weeks=i) for i in range(num_weeks)]

    # Build the work queue: for each job, remaining hours per future phase
    job_queue = []
    for job_key, analysis in job_analyses.items():
        if analysis is None:
            continue
        current_phase = analysis['current_phase']
        hit_current = False
        remaining_phases = []
        for phase in PHASE_NAMES:
            remaining = analysis['phase_remaining'][phase]
            if phase == current_phase:
                hit_current = True
            if hit_current and remaining > 0:
                remaining_phases.append({'phase': phase, 'remaining': remaining})
        if remaining_phases:
            job_queue.append({
                'job': job_key,
                'phases': remaining_phases,
                'phase_idx': 0,  # index into remaining_phases
                'promised': analysis.get('promised_date'),
                'pct_complete': analysis.get('pct_complete', 0),
            })

    # Sort by priority: earliest promised date first, then most-complete first
    def sort_key(j):
        promised = j['promised']
        if promised is None or (isinstance(promised, float)):
            date_key = datetime(2099, 12, 31)
        elif isinstance(promised, str):
            try:
                date_key = datetime.strptime(promised, '%Y-%m-%d')
            except ValueError:
                date_key = datetime(2099, 12, 31)
        else:
            date_key = promised
        return (date_key, -j['pct_complete'])

    job_queue.sort(key=sort_key)

    # Result: job -> week -> phase
    schedule = {j['job']: {} for j in job_queue}

    # Simulate week by week
    for week in week_dates:
        # Budget available this week per phase
        budget = {p: PHASE_CAPACITY_PER_WEEK[p] for p in PHASE_NAMES}

        # Allocate capacity to jobs in priority order
        for j in job_queue:
            if j['phase_idx'] >= len(j['phases']):
                continue  # job is done

            current = j['phases'][j['phase_idx']]
            phase = current['phase']
            remaining = current['remaining']

            if budget[phase] <= 0:
                # No capacity left in this phase this week — job waits
                # Still mark it as waiting in this phase so the Gantt shows it queued
                schedule[j['job']][week] = phase + ' *'
                continue

            # Allocate hours (up to what's available)
            hours_this_week = min(remaining, budget[phase])
            budget[phase] -= hours_this_week
            current['remaining'] -= hours_this_week

            schedule[j['job']][week] = phase

            # If this phase is done, advance to next phase
            if current['remaining'] <= 0:
                j['phase_idx'] += 1

    return schedule


def generate_week_columns(start_date, num_weeks=24):
    """Generate Monday-based week column dates."""
    start = start_date - timedelta(days=start_date.weekday())
    return [start + timedelta(weeks=i) for i in range(num_weeks)]


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    jobs = pd.read_csv(os.path.join(DATA_DIR, 'jobs.csv'),
                       parse_dates=['Order_Date'],
                       low_memory=False)
    ops = pd.read_csv(os.path.join(DATA_DIR, 'job_operations.csv'), low_memory=False)
    customers = pd.read_csv(os.path.join(DATA_DIR, 'customers.csv'))
    deliveries = pd.read_csv(os.path.join(DATA_DIR, 'deliveries.csv'),
                             parse_dates=['Promised_Date', 'Shipped_Date'])

    cust_map = dict(zip(customers['Customer'], customers['Name']))

    # Active jobs only — exclude jobs with no operations
    active = jobs[jobs['Status'].isin(['Active', 'Released', 'Hold'])].copy()
    jobs_with_ops = ops.groupby('Job')['Sequence'].count()
    active = active[active['Job'].isin(jobs_with_ops.index)]

    today = datetime.now()
    weeks = generate_week_columns(today, num_weeks=24)

    # ---- Phase 1: Analyze every job and collect promised dates ----
    job_data = {}   # job_num -> {analysis, promised, job_row}
    for _, job in active.iterrows():
        job_num = job['Job']
        job_ops = ops[ops['Job'] == job_num].copy()
        analysis = analyze_job(job_ops)
        if analysis is None:
            continue

        job_del = deliveries[deliveries['Job'] == job_num]
        promised = None
        if not job_del.empty and job_del['Promised_Date'].notna().any():
            promised = job_del['Promised_Date'].dropna().iloc[0]

        # Attach promised date to analysis so the scheduler can prioritize
        analysis['promised_date'] = promised
        job_data[job_num] = {'analysis': analysis, 'promised': promised, 'job_row': job}

    # ---- Phase 2: Run the capacity-constrained scheduler across ALL jobs ----
    all_analyses = {jn: jd['analysis'] for jn, jd in job_data.items()}
    global_schedule = schedule_all_jobs(all_analyses, today, num_weeks=24)

    # ---- Phase 3: Build output rows using the global schedule ----
    rows = []
    for job_num, jd in job_data.items():
        analysis = jd['analysis']
        promised = jd['promised']
        job = jd['job_row']
        job_sched = global_schedule.get(job_num, {})

        promised_str = ''
        if promised is not None and pd.notna(promised):
            promised_str = promised.strftime('%Y-%m-%d') if hasattr(promised, 'strftime') else str(promised)

        row = {
            'Customer': cust_map.get(job['Customer'], job['Customer']),
            'Job #': job_num,
            'Description': job.get('Description', ''),
            'Total $': job.get('Total_Price', 0),
            'PO Recvd': job['Order_Date'].strftime('%Y-%m-%d') if pd.notna(job['Order_Date']) else '',
            'Promised Date': promised_str,
            'Current Phase': analysis['current_phase'],
            '% Complete': round(analysis['pct_complete'], 1),
            'Est Hrs': round(analysis['total_est_hrs'], 0),
            'Remaining Hrs': round(analysis['total_remaining_hrs'], 0),
        }

        for week_start in weeks:
            col_name = week_start.strftime('%m/%d')
            phase_val = job_sched.get(week_start, '')
            row[col_name] = phase_val

        rows.append(row)

    df = pd.DataFrame(rows)

    # Sort by remaining hours descending (biggest jobs first)
    df = df.sort_values('Remaining Hrs', ascending=False).reset_index(drop=True)

    # Write to Excel with formatting
    outfile = os.path.join(OUT_DIR, 'Master_Schedule.xlsx')
    with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule', index=False)

        ws = writer.sheets['Schedule']

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 18)

        # Color-code phase cells
        from openpyxl.styles import PatternFill
        phase_colors = {
            'PRE-FAB':     'FFF2CC',   # light yellow
            'FABRICATION': 'D9E2F3',   # light blue
            'TESTING':     'E2EFDA',   # light green
            'CLEANING':    'FCE4D6',   # light orange
            'PAINT':       'E4DFEC',   # light purple
            'SHIPPING':    'D6DCE4',   # light gray
        }

        # Find where week columns start
        header_row = [cell.value for cell in ws[1]]
        week_col_start = None
        for i, h in enumerate(header_row):
            if h and '/' in str(h) and len(str(h)) <= 5:
                week_col_start = i + 1  # 1-indexed
                break

        # Lighter colors for queued (waiting) weeks
        queue_colors = {
            'PRE-FAB':      'FFF9E6',
            'FABRICATION':  'EDF2FB',
            'TESTING':      'F2F9EE',
            'CLEANING':     'FEF3EC',
            'PAINT':        'F3F0F8',
            'SHIPPING':     'ECEEF2',
        }

        if week_col_start:
            for row_cells in ws.iter_rows(min_row=2, min_col=week_col_start,
                                          max_col=ws.max_column):
                for cell in row_cells:
                    val = str(cell.value or '')
                    if val in phase_colors:
                        cell.fill = PatternFill(start_color=phase_colors[val],
                                                end_color=phase_colors[val],
                                                fill_type='solid')
                    elif val.endswith(' *'):
                        base_phase = val.replace(' *', '')
                        if base_phase in queue_colors:
                            cell.fill = PatternFill(start_color=queue_colors[base_phase],
                                                    end_color=queue_colors[base_phase],
                                                    fill_type='solid')

    print(f"Generated: {outfile}")
    print(f"  {len(df)} active jobs across {len(weeks)} weeks")
    print(f"  Week range: {weeks[0].strftime('%m/%d/%Y')} to {weeks[-1].strftime('%m/%d/%Y')}")
    print(f"  Phase capacity (hrs/week shared across all jobs):")
    for phase, cap in PHASE_CAPACITY_PER_WEEK.items():
        print(f"    {phase:14s} {cap:>4} hrs/week")
    print()

    # Summary stats
    in_prog = df[df['% Complete'].between(0.1, 99.9)]
    print(f"  In-progress jobs: {len(in_prog)}")
    print(f"  Not started:      {len(df[df['% Complete'] == 0])}")
    print(f"  Total remaining:  {df['Remaining Hrs'].sum():,.0f} hrs")
    print()
    print("Sample (top 10 by remaining hours):")
    sample_cols = ['Customer', 'Job #', 'Current Phase', '% Complete', 'Remaining Hrs']
    print(df.head(10)[sample_cols].to_string(index=False))


if __name__ == '__main__':
    main()
