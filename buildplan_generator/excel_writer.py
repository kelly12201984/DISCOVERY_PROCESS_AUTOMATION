"""
Generate Excel build plans in Savannah Tank's format.
Mirrors the operation order from their master template.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# Section order matches Savannah Tank's master build plan
SECTION_ORDER = [
    ("Receive & Inspect", [".010", ".QC1"]),
    ("Nozzles and Manways", [
        ".100", ".105", ".107", ".110", ".115", ".120", ".125",
        ".130", ".135", ".140", ".145", ".147"
    ]),
    ("Cut & Prep Parts", [".150", ".155", ".156", ".160"]),
    ("Supports", [".165", ".170", ".175", ".180", ".185"]),
    ("Misc. Parts", [
        ".200", ".205", ".210", ".215", ".220", ".225", ".230",
        ".235", ".240", ".245", ".250", ".252", ".255", ".257",
        ".260", ".261"
    ]),
    ("Ladders, Handrails and Platforms", [
        ".262", ".265", ".270", ".272", ".275", ".280", ".282",
        ".285", ".290", ".292", ".295", ".QC2"
    ]),
    ("Shells", [".300", ".QC3", ".305", ".310", ".315", ".320", ".325", ".326"]),
    ("Top Head", [
        ".350", ".QC4", ".355", ".360", ".365", ".366", ".370",
        ".375", ".380", ".382", ".383", ".385", ".390", ".395"
    ]),
    ("Bottom Head", [
        ".450", ".QC4", ".455", ".460", ".465", ".466", ".470",
        ".475", ".477", ".478", ".479", ".482", ".483", ".485",
        ".490", ".495"
    ]),
    ("Skirt", [".500", ".510", ".515", ".520", ".530", ".540", ".545"]),
    ("Internal Coils", [".580", ".590"]),
    ("Build Main Tank and Install Internals", [
        ".OI1", ".600", ".605", ".610", ".615", ".616", ".620",
        ".625", ".630", ".635", ".637", ".638", ".QC5",
        ".640", ".641", ".645", ".800"
    ]),
    ("Install Nozzles and Externals", [
        ".700", ".705", ".710", ".711", ".715", ".717", ".720",
        ".721", ".725", ".726", ".730", ".735", ".740", ".745",
        ".750", ".755", ".760", ".QC51", ".765"
    ]),
    ("Inspect and Repair", [".770", ".QC6", ".OI4", ".OI2", ".850", ".855", ".860"]),
    ("Test and Clean", [
        ".801", ".900", ".910", ".QC7", ".OI3", ".920", ".QC8",
        ".802", ".QC9", ".QC90"
    ]),
    ("Ship", [".940", ".950", ".QC95", ".960", ".970", ".971", ".972", ".973", ".974", ".975"]),
]


# Standard colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
NOTE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HIGH_CONF_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
MED_CONF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LOW_CONF_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CONF_FILLS = {
    "high": HIGH_CONF_FILL,
    "medium": MED_CONF_FILL,
    "low": LOW_CONF_FILL,
    "none": LOW_CONF_FILL,
}


def write_build_plan(predictions: pd.DataFrame, job_no: str, job_desc: str,
                     sibling_jobs: list[str], output_path: str | Path,
                     total_est: float = None) -> Path:
    """Write a build plan Excel file.

    Args:
        predictions: DataFrame from HoursPredictor with columns:
            op_code, work_center, description, predicted_hours, confidence,
            sibling_range, source_jobs, notes
        job_no: Target job number.
        job_desc: Job description (tank name).
        sibling_jobs: List of similar job numbers used.
        output_path: Where to save the .xlsx.
        total_est: Optional total estimated hours override.

    Returns:
        Path to the saved file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Build Plan with Hours ---
    ws = wb.active
    ws.title = "Build Plan w Hours"
    _write_plan_sheet(ws, predictions, job_no, job_desc, sibling_jobs,
                      include_hours=True, include_confidence=True, total_est=total_est)

    # --- Sheet 2: Build Plan without Hours (shop floor version) ---
    ws2 = wb.create_sheet("Build Plan - No Hours")
    _write_plan_sheet(ws2, predictions, job_no, job_desc, sibling_jobs,
                      include_hours=False, include_confidence=False, total_est=total_est)

    # --- Sheet 3: Sibling Analysis ---
    ws3 = wb.create_sheet("Sibling Analysis")
    _write_sibling_sheet(ws3, predictions, sibling_jobs)

    wb.save(output_path)
    return output_path


def _write_plan_sheet(ws, predictions, job_no, job_desc, sibling_jobs,
                      include_hours, include_confidence, total_est):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_font = Font(bold=True, size=11, color="1F4E79")
    title_font = Font(bold=True, size=14)

    # Title rows
    ws["A1"] = f"Job #:"
    ws["A1"].font = Font(bold=True, size=12)
    ws["B1"] = job_no
    ws["B1"].font = title_font
    ws["C1"] = "Description:"
    ws["C1"].font = Font(bold=True)
    ws["D1"] = job_desc

    ws["A2"] = "Similar jobs used:"
    ws["B2"] = ", ".join(sibling_jobs[:5])
    ws["B2"].font = Font(italic=True, color="666666")

    if total_est:
        ws["F1"] = "Total Est Hours:"
        ws["G1"] = total_est
        ws["G1"].font = Font(bold=True, size=12)

    # Column headers
    row = 4
    headers = ["Work Center", "", "Op #", "Task", "Est Hours"]
    if include_hours:
        headers += ["Predicted Hours"]
    if include_confidence:
        headers += ["Confidence", "Range"]
    headers += ["Notes", "Completed By", "Date"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Build a lookup from op_code to prediction row
    pred_lookup = {}
    for _, p in predictions.iterrows():
        code = str(p.get("op_code", "")).strip()
        pred_lookup[code] = p

    # Also build lookup for ops not in the standard template
    template_ops = set()
    for _, op_codes in SECTION_ORDER:
        template_ops.update(op_codes)

    extra_ops = [p for code, p in pred_lookup.items() if code not in template_ops]

    row = 5
    total_predicted = 0
    written_ops = set()

    for section_name, op_codes in SECTION_ORDER:
        # Check if any ops in this section are present AND not already written
        section_ops = [pred_lookup[code] for code in op_codes
                       if code in pred_lookup and code not in written_ops]
        if not section_ops:
            continue

        # Section header
        ws.cell(row=row, column=4, value=section_name).font = section_font
        for col_idx in range(1, 12):
            ws.cell(row=row, column=col_idx).fill = SECTION_FILL
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

        for code in op_codes:
            if code not in pred_lookup or code in written_ops:
                continue
            written_ops.add(code)
            p = pred_lookup[code]
            row = _write_op_row(ws, row, p, include_hours, include_confidence)
            if include_hours:
                total_predicted += float(p.get("predicted_hours", 0))

    # Extra ops not in template
    if extra_ops:
        ws.cell(row=row, column=4, value="Additional Operations").font = section_font
        for col_idx in range(1, 12):
            ws.cell(row=row, column=col_idx).fill = SECTION_FILL
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

        for p in extra_ops:
            row = _write_op_row(ws, row, p, include_hours, include_confidence)
            if include_hours:
                total_predicted += float(p.get("predicted_hours", 0))

    # Total row
    row += 1
    ws.cell(row=row, column=4, value="Total Predicted Hours").font = Font(bold=True, size=12)
    if include_hours:
        ws.cell(row=row, column=6, value=round(total_predicted, 1)).font = Font(bold=True, size=12)

    # Column widths
    widths = [14, 4, 8, 55, 12, 14, 12, 20, 30, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_op_row(ws, row, p, include_hours, include_confidence) -> int:
    wc = str(p.get("work_center", "")).upper()
    op_code = str(p.get("op_code", ""))
    desc = str(p.get("description", ""))
    est_hrs = p.get("est_hours", 0)
    predicted = p.get("predicted_hours", 0)
    confidence = str(p.get("confidence", "none"))
    hr_range = str(p.get("sibling_range", ""))
    notes = str(p.get("notes", ""))
    if notes == "nan":
        notes = ""

    ws.cell(row=row, column=1, value=wc)
    ws.cell(row=row, column=3, value=op_code)
    ws.cell(row=row, column=4, value=desc)

    if est_hrs and float(est_hrs) > 0:
        ws.cell(row=row, column=5, value=float(est_hrs))

    if include_hours and predicted and float(predicted) > 0:
        cell = ws.cell(row=row, column=6, value=float(predicted))
        cell.fill = CONF_FILLS.get(confidence, LOW_CONF_FILL)

    if include_confidence:
        ws.cell(row=row, column=7, value=confidence)
        ws.cell(row=row, column=8, value=hr_range)

    if notes:
        note_cell = ws.cell(row=row, column=9, value=notes)
        note_cell.fill = NOTE_FILL

    for col_idx in range(1, 12):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

    return row + 1


def _write_sibling_sheet(ws, predictions, sibling_jobs):
    """Reference sheet showing which historical jobs were used."""
    ws["A1"] = "Sibling Job Analysis"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Jobs used for predictions:"
    for i, job in enumerate(sibling_jobs):
        ws.cell(row=4 + i, column=1, value=job)

    ws["A7" if len(sibling_jobs) < 4 else f"A{5 + len(sibling_jobs) + 1}"] = "Per-Operation Source Detail"
    start_row = 8 if len(sibling_jobs) < 4 else 6 + len(sibling_jobs) + 1
    headers = ["Op Code", "Description", "Predicted Hrs", "Confidence", "Range", "Source Jobs"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)

    for i, (_, p) in enumerate(predictions.iterrows(), start_row + 1):
        ws.cell(row=i, column=1, value=str(p.get("op_code", "")))
        ws.cell(row=i, column=2, value=str(p.get("description", "")))
        ws.cell(row=i, column=3, value=float(p.get("predicted_hours", 0)))
        ws.cell(row=i, column=4, value=str(p.get("confidence", "")))
        ws.cell(row=i, column=5, value=str(p.get("sibling_range", "")))
        ws.cell(row=i, column=6, value=str(p.get("source_jobs", "")))

    for w, col in zip([12, 50, 14, 12, 20, 40], range(1, 7)):
        ws.column_dimensions[get_column_letter(col)].width = w
