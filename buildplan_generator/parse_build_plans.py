"""
Parse all Excel build plans into a normalized table.
Handles both formats:
  - Old: "NNN-YY Build Plan.xlsx" (cols B/C/D/E/G, sheet "New Build Plan")
  - New: "YY-NNN MASTER BUILD PLAN.xlsx" (cols A/C/D/E, sheet "Build Plan w Hours")

Output: one row per operation per job:
  job_no, op_code, work_center, description, est_hours, section, notes
"""
import re
import openpyxl
from pathlib import Path
import pandas as pd
from config import BUILD_PLANS_DIR, PARSED_BUILD_PLANS, DATA_DIR, WC_MAP


def extract_job_number(filename: str) -> str:
    """Extract job number from filename, normalize to SEQ-YY format."""
    name = Path(filename).stem

    # "YY-NNN MASTER BUILD PLAN" format
    m = re.match(r"(\d{2})-(\d{3})", name)
    if m:
        return f"{m.group(2)}-{m.group(1)}"

    # "NNN-YY Build Plan" format
    m = re.match(r"(\d{3})-(\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    return None


def safe_cell_value(cell):
    """Get cell value, handling merged cells."""
    if hasattr(cell, "value"):
        return cell.value
    return None


def safe_column_letter(cell):
    """Get column letter, handling merged cells."""
    if hasattr(cell, "column_letter"):
        return cell.column_letter
    return None


def parse_old_format(ws) -> list[dict]:
    """Parse 'New Build Plan' sheet: B=wc, C=op#, D=desc, E=hrs, G=notes."""
    rows = []
    current_section = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        vals = {}
        for c in row:
            letter = safe_column_letter(c)
            val = safe_cell_value(c)
            if letter and val is not None:
                vals[letter] = val

        # Skip header rows
        if vals.get("B") in ("Job #", "Work Center"):
            continue

        # Section header: D has value, B is empty, E is empty
        if "D" in vals and "B" not in vals and "E" not in vals:
            d = str(vals["D"]).strip()
            if d and d not in ("Description", "Total Estimate Hours", "Job Estimated Hours",
                               "Fill in Green Cells Only as Needed"):
                current_section = d
            continue

        # Operation row: has op code in C and work center in B
        if "C" in vals and "B" in vals:
            wc_raw = str(vals["B"]).strip().lower()
            wc = WC_MAP.get(wc_raw, wc_raw.upper())
            op_code = str(vals["C"]).strip()
            if not op_code.startswith("."):
                op_code = "." + op_code
            desc = str(vals.get("D", "")).strip()
            est_hrs = vals.get("E", 0)
            if not isinstance(est_hrs, (int, float)):
                est_hrs = 0
            notes = str(vals.get("G", "")).strip() if "G" in vals else ""
            if notes in ("Fill in Green Cells Only as Needed", "Notes to Add"):
                notes = ""

            rows.append({
                "op_code": op_code,
                "work_center": wc,
                "description": desc,
                "est_hours": float(est_hrs),
                "section": current_section,
                "notes": notes,
            })

    return rows


def parse_new_format(ws) -> list[dict]:
    """Parse 'Build Plan w Hours' sheet: A=wc, C=op#, D=task, E=hrs.
    Notes are embedded in task description after _x000D_ newline."""
    rows = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        vals = {}
        for c in row:
            letter = safe_column_letter(c)
            val = safe_cell_value(c)
            if letter and val is not None:
                vals[letter] = val

        # Skip headers and metadata
        if vals.get("A") in ("Job #:", "Work Center", None) and "C" not in vals:
            continue

        # Operation row: has op code in C
        op_val = vals.get("C")
        if op_val is None or op_val == "Op #":
            continue

        op_code = str(op_val).strip()
        if op_code == "0":
            continue
        if not op_code.startswith("."):
            op_code = "." + op_code

        wc_raw = str(vals.get("A", "")).strip()
        if wc_raw == "0":
            wc_raw = ""
        wc = WC_MAP.get(wc_raw.lower(), wc_raw.upper()) if wc_raw else ""

        # Description may contain inline notes after _x000D_\n
        raw_desc = str(vals.get("D", ""))
        if raw_desc == "0":
            raw_desc = ""
        parts = re.split(r"_x000D_\n|_x000D_\\n|\n", raw_desc, maxsplit=1)
        desc = parts[0].strip()
        notes = parts[1].strip() if len(parts) > 1 else ""

        est_hrs = vals.get("E", 0)
        if not isinstance(est_hrs, (int, float)):
            est_hrs = 0

        rows.append({
            "op_code": op_code,
            "work_center": wc,
            "description": desc,
            "est_hours": float(est_hrs),
            "section": "",  # new format doesn't use section headers
            "notes": notes,
        })

    return rows


def parse_all_plans() -> pd.DataFrame:
    all_rows = []
    files = sorted(BUILD_PLANS_DIR.glob("*.xlsx"))
    parsed = 0
    errors = []

    for f in files:
        if f.name.startswith("~") or f.name == "MASTER BUILD PLAN.xlsx":
            continue

        job_no = extract_job_number(f.name)
        if not job_no:
            errors.append(f"  Skip (no job#): {f.name}")
            continue

        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
            sheets = wb.sheetnames

            if "Build Plan w Hours" in sheets:
                ops = parse_new_format(wb["Build Plan w Hours"])
            elif "New Build Plan" in sheets:
                ops = parse_old_format(wb["New Build Plan"])
            else:
                ops = parse_old_format(wb[sheets[0]])

            for op in ops:
                op["job_no"] = job_no
            all_rows.extend(ops)
            parsed += 1
            wb.close()
        except Exception as e:
            errors.append(f"  Error {f.name}: {e}")

    print(f"Parsed {parsed} build plans, {len(all_rows)} total operations")
    if errors:
        print(f"  {len(errors)} issues:")
        for e in errors[:10]:
            print(e)

    return pd.DataFrame(all_rows)


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("Parsing all build plans...")
    df = parse_all_plans()

    if len(df) > 0:
        # Stats
        print(f"\nUnique jobs: {df['job_no'].nunique()}")
        print(f"Unique op codes: {df['op_code'].nunique()}")
        print(f"Ops with hours > 0: {(df['est_hours'] > 0).sum()}")
        print(f"Ops with notes: {(df['notes'].str.len() > 0).sum()}")
        print(f"\nTop work centers:")
        print(df["work_center"].value_counts().head(10).to_string())
        print(f"\nTop op codes (by frequency):")
        print(df["op_code"].value_counts().head(20).to_string())

        df.to_csv(PARSED_BUILD_PLANS, index=False)
        print(f"\nSaved: {PARSED_BUILD_PLANS}")


if __name__ == "__main__":
    main()
